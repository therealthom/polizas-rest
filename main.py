import io
import os
import tempfile
from flask import Flask, request, jsonify, make_response
import requests
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from google.cloud import storage, bigquery
from openpyxl import load_workbook
from flask import Flask, send_file


app = Flask(__name__)
CORS(app)

# Configura la conexión al bucket de Google Cloud Storage
bucket_name = 'testing-polizas'
storage_client = storage.Client()
bucket = storage_client.bucket(bucket_name)
origen = 'pending'


@app.route('/descargar-datos', methods=['GET'])
def descargar_datos():
    # Configurar el proyecto y el ID del conjunto de datos
    project_id = 'demoasf'
    dataset_id = 'polizas'

    # Configurar el nombre de la tabla
    table_name = 'poliza_egreso'

    try:
        # Consultar la tabla en BigQuery
        client = bigquery.Client(project=project_id)
        query = f'SELECT * FROM `{project_id}.{dataset_id}.{table_name}`'
        df = client.query(query).to_dataframe()

        # Generar el archivo XLSX en memoria
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)
        workbook.save(output)
        output.seek(0)

        # Crear la respuesta con el archivo adjunto
        response = make_response(output.getvalue())
        response.headers.set('Content-Disposition', 'attachment', filename='datos.xlsx')
        response.headers.set('Access-Control-Allow-Origin', '*')
        response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return response
    except Exception as e:
        return jsonify({'mensaje': 'Error al descargar los datos: ' + str(e)}), 500


@app.route('/consultar-tabla', methods=['GET'])
def consultar_tabla():
    # Configurar el proyecto y el ID del conjunto de datos
    project_id = 'demoasf'
    dataset_id = 'polizas'

    # Configurar el nombre de la tabla
    table_name = 'poliza_egreso'

    respuesta_headers = {
        'Access-Control-Allow-Origin': '*'
    }
    try:
        # Consultar la tabla en BigQuery utilizando el cliente predeterminado
        client = bigquery.Client(project=project_id)
        query = f'SELECT fecha, codigo_contable, codigo_presupuestal, no_asiento, ejercicio, debe, haber FROM `{project_id}.{dataset_id}.{table_name}`'
        job = client.query(query)
        result = job.result()

        # Convertir el resultado en una lista de diccionarios
        rows = [dict(row.items()) for row in result]

        return jsonify(rows), 200, respuesta_headers
    except Exception as e:
        return jsonify({'mensaje': 'Error al consultar la tabla: ' + str(e)}), 500, respuesta_headers


# Ruta del servicio REST que recibe los archivos PDF como parámetros
@app.route('/guardar-archivos', methods=['POST'])
def guardar_archivos():
    # Verificar si se enviaron archivos
    if 'archivos' not in request.files:
        return jsonify({'mensaje': 'No se enviaron archivos.'}), 400

    # Obtén los archivos PDF enviados en la solicitud
    archivos = request.files.getlist('archivos')

    # Verificar si no se enviaron archivos vacíos
    if len(archivos) == 0:
        return jsonify({'mensaje': 'No se enviaron archivos válidos.'}), 400

    # Crea una carpeta temporal para almacenar los archivos
    carpeta_temporal = tempfile.mkdtemp()

    # Guarda cada archivo en la carpeta temporal
    exito = True
    for archivo in archivos:
        archivo_path = os.path.join(carpeta_temporal, archivo.filename)
        archivo.save(archivo_path)

        # Sube el archivo al bucket de Google Cloud Storage en la carpeta "pending"
        blob = bucket.blob('pending/{}'.format(archivo.filename))
        try:
            blob.upload_from_filename(archivo_path)
        except Exception as e:
            exito = False

        # Elimina el archivo temporal
        os.remove(archivo_path)

    respuesta_headers = {
        'Access-Control-Allow-Origin': '*'
    }
    # Retorna un JSON con el resultado de la operación
    if exito:
        return jsonify({'exito': True, 'mensaje': 'Archivos guardados correctamente'}), 200, respuesta_headers
    else:
        return jsonify(
            {'exito': False, 'mensaje': 'Error al guardar archivos en Google Cloud Storage'}), 200, respuesta_headers


@app.route('/procesar-archivos', methods=['GET'])
def procesar_archivos():
    function_url = 'https://us-central1-demoasf.cloudfunctions.net/function-polizasegresos'
    headers = {
        'Content-Type': 'application/json'
    }
    data = {}
    response = requests.get(
        url=function_url,
        headers=headers,
        params=data
    )
    respuesta_headers = {
        'Access-Control-Allow-Origin': '*',
        'Content-Type': 'application/json'
    }
    return jsonify(response.json()), 200, respuesta_headers


@app.route('/listar-archivos', methods=['GET'])
def listar_archivos():

    try:
        # Obtener lista de archivos en el bucket
        blobs = bucket.list_blobs(prefix=f"{origen}/")

        # Crear una lista con los nombres de los archivos
        pdf_files = [blob.name.split("/")[-1] for blob in blobs if blob.name.lower().endswith('.pdf')]

        return jsonify({'archivos': pdf_files}), 200
    except Exception as e:
        return jsonify({'mensaje': 'Error al obtener la lista de archivos: ' + str(e)}), 500


@app.route('/modify-excel', methods=['GET'])
def modify_excel():
    # Get environment variables
    bucket_name = "testing-polizas"
    blob_name = "plantilla-poliza-egreso.xlsx"
    dataset_id = "polizas"
    table_id = "poliza_egreso"
    project_id = "demoasf"

    # Initialize a Google Cloud Storage client
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(blob_name)

    # Download the XLSX file from GCS to local storage
    blob.download_to_filename('/tmp/input.xlsx')

    # Initialize a BigQuery client
    bigquery_client = bigquery.Client()
    table = bigquery_client.dataset(dataset_id).table(table_id)

    # Fetch data from BigQuery
    query = f'SELECT * FROM `{project_id}.{dataset_id}.{table_id}`'
    query_job = bigquery_client.query(query)
    results = query_job.result()

    # Load the workbook and select the active sheet
    wb = load_workbook('/tmp/input.xlsx')
    ws = wb.active

    # Find the first empty row in the sheet
    first_empty_row = next((i for i, row in enumerate(ws.iter_rows(), start=1) if all(cell.value is None for cell in row)), None)

    # Modify the XLSX file with the fetched data
    for row in results:
        for col_num, value in enumerate(row, start=1):
            ws.cell(row=first_empty_row, column=col_num, value=value)
        first_empty_row += 1

    # Save the modified XLSX file
    wb.save('/tmp/output.xlsx')

    # Upload the modified XLSX file back to GCS
    blob.upload_from_filename('/tmp/output.xlsx')

    # Return the XLSX file for download
    return send_file('/tmp/output.xlsx', as_attachment=True)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080, debug=True)
